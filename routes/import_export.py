import io
import re
from datetime import date, datetime
from flask import (Blueprint, render_template, request,
                   redirect, url_for, flash, session, send_file)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from models import db, Expense
from utils import login_required
from routes.expenses import DEFAULT_CATEGORIES, MONTH_NAMES

import_export_bp = Blueprint('import_export', __name__)

# Fast lookup for category normalisation
_STANDARD_CATS_LOWER = {c.lower(): c for c in DEFAULT_CATEGORIES}


def _normalize_category(raw: str) -> str:
    """Return the matching standard category, or 'Miscellaneous' if unknown."""
    if not raw:
        return 'Miscellaneous'
    return _STANDARD_CATS_LOWER.get(raw.strip().lower(), 'Miscellaneous')

# ------------------------------------------------------------------ #
# Constants                                                           #
# ------------------------------------------------------------------ #

_HEADERS     = ['Sr. No.', 'Date', 'Title', 'Description', 'Category', 'Mode', 'Amount', 'Split']
_COL_WIDTHS  = [8, 14, 28, 34, 16, 18, 12, 12]

_MONTH_ABBR  = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
_MONTH_MAP   = {m: i + 1 for i, m in enumerate(_MONTH_ABBR)}

# Mode strings that indicate a friend (not the user) paid
_FRIEND_TOKENS = {'friend', 'aayush', 'kunal'}   # extend as needed

# ------------------------------------------------------------------ #
# Style helpers                                                       #
# ------------------------------------------------------------------ #

_ACCENT      = '1A472A'
_HEADER_BG   = PatternFill('solid', fgColor=_ACCENT)
_HEADER_FG   = Font(bold=True, color='FFFFFF', size=11)
_SIDE        = Side(style='thin', color='E0DDD6')
_BORDER      = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
_ALT_FILL    = PatternFill('solid', fgColor='F5F4F1')


def _style_ws(ws, ncols, widths):
    for col in range(1, ncols + 1):
        c = ws.cell(1, col)
        c.fill, c.font = _HEADER_BG, _HEADER_FG
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = _BORDER
    ws.row_dimensions[1].height = 22
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'


def _write_data_row(ws, row_idx, values, alt=False):
    fill = _ALT_FILL if alt else None
    for col, val in enumerate(values, 1):
        c = ws.cell(row_idx, col, val)
        c.alignment = Alignment(vertical='center')
        c.border = _BORDER
        if fill:
            c.fill = fill
    ws.cell(row_idx, 7).number_format = '#,##0.00'
    ws.cell(row_idx, 8).number_format = '#,##0.00'


# ------------------------------------------------------------------ #
# Sheet-name utilities                                                #
# ------------------------------------------------------------------ #

def _tab_name(year: int, month: int) -> str:
    return f'{_MONTH_ABBR[month - 1]} {str(year)[-2:]}'


def _parse_tab_name(name: str):
    """Parse 'Jun 26' → (2026, 6). Returns None if unrecognised."""
    m = re.fullmatch(r'([A-Za-z]{3})\s*(\d{2,4})', name.strip())
    if not m:
        return None
    mon, yr = m.group(1).capitalize(), int(m.group(2))
    month = _MONTH_MAP.get(mon)
    if not month:
        return None
    year = yr + 2000 if yr < 100 else yr
    return year, month


# ------------------------------------------------------------------ #
# paid_by_user inference                                              #
# ------------------------------------------------------------------ #

def _infer_paid_by_user(mode, who_paid=None):
    """Return True if the user paid, False if a friend paid."""
    if who_paid is not None:
        return str(who_paid).strip().upper() == 'I'
    if not mode:
        return True
    m = str(mode).lower().strip()
    # "Friend Paid", "Aayush Paid", "FRIEND", etc.
    if any(tok in m for tok in _FRIEND_TOKENS) or m.endswith('paid'):
        return False
    return True


# ------------------------------------------------------------------ #
# Import: layout detection                                            #
# ------------------------------------------------------------------ #

def _detect_layout(raw_headers):
    """Map column names to 0-based indices for the three known layouts."""
    hdrs = [str(h).strip().lower() if h else '' for h in raw_headers]

    def find(*candidates):
        for cand in candidates:
            try:
                return hdrs.index(cand.lower())
            except ValueError:
                pass
        return -1

    title_col    = find('title', 'name', 'brand/title')
    date_col     = find('date')
    desc_col     = find('description')
    cat_col      = find('category')
    mode_col     = find('mode')
    who_paid_col = find('who paid?')
    amount_col   = find('amount')
    split_col    = find('split')
    paid_col     = find('paid')     # Aug–Nov 25 layout

    if who_paid_col >= 0:
        # Dec 25 layout: Who Paid? col present; Split is separate from Amount
        return {
            'date': date_col, 'title': title_col, 'desc': desc_col,
            'cat': cat_col, 'mode': mode_col, 'amount': amount_col,
            'split': split_col, 'who_paid': who_paid_col,
        }

    if paid_col >= 0 and split_col < 0:
        # Aug–Nov 25 layout: "Paid" column is the split
        return {
            'date': date_col, 'title': title_col, 'desc': desc_col,
            'cat': cat_col, 'mode': mode_col, 'amount': amount_col,
            'split': paid_col, 'who_paid': -1,
        }

    # Jan 26+ standard layout
    return {
        'date': date_col, 'title': title_col, 'desc': desc_col,
        'cat': cat_col, 'mode': mode_col, 'amount': amount_col,
        'split': split_col, 'who_paid': -1,
    }


def _cell(row, idx):
    """Safe cell access from a tuple row."""
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ------------------------------------------------------------------ #
# Export workbook builder                                             #
# ------------------------------------------------------------------ #

def _build_export_wb(expenses_by_month: dict) -> Workbook:
    """Build one workbook with one sheet per (year, month) key."""
    wb = Workbook()
    wb.remove(wb.active)          # remove default blank sheet

    for (year, month), rows in sorted(expenses_by_month.items()):
        ws = wb.create_sheet(title=_tab_name(year, month))
        for col, h in enumerate(_HEADERS, 1):
            ws.cell(1, col, h)
        _style_ws(ws, len(_HEADERS), _COL_WIDTHS)

        for i, e in enumerate(rows):
            split_val = e.split if e.split is not None else e.amount
            values = [
                i + 1,
                e.date.strftime('%Y-%m-%d'),
                e.title,
                e.description or '',
                e.category,
                e.mode or '',
                round(e.amount, 2),
                round(split_val, 2),
            ]
            _write_data_row(ws, i + 2, values, alt=(i % 2 == 1))

    return wb


# ------------------------------------------------------------------ #
# Template workbook builder                                           #
# ------------------------------------------------------------------ #

def _build_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mon YY'

    for col, h in enumerate(_HEADERS, 1):
        ws.cell(1, col, h)
    _style_ws(ws, len(_HEADERS), _COL_WIDTHS)

    example_fill = PatternFill('solid', fgColor='EDEBE6')
    example_font = Font(color='999999', italic=True, size=10)
    examples = [
        [1, '2026-06-01', 'STARBUCKS',  'Iced coffee',      'Eating out', 'AMEX',        6.50,  6.50],
        [2, '2026-06-05', 'WALMART',    'Groceries',         'Grocery',    'FRIEND',      42.00, 10.50],
        [3, '2026-06-10', 'NETFLIX',    'Monthly sub',       'Bills',      'Chase Visa',  17.99, 17.99],
        [4, '2026-06-15', 'DOORDASH',   'Pizza with Aayush', 'Eating out', 'Aayush Paid', 38.00, 19.00],
    ]
    for i, row in enumerate(examples):
        for col, val in enumerate(row, 1):
            c = ws.cell(i + 2, col, val)
            c.fill, c.font = example_fill, example_font
            c.alignment = Alignment(vertical='center')
            c.border = _BORDER
        ws.cell(i + 2, 7).number_format = '#,##0.00'
        ws.cell(i + 2, 8).number_format = '#,##0.00'

    cat_formula = '"%s"' % ','.join(DEFAULT_CATEGORIES)
    dv = DataValidation(type='list', formula1=cat_formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.sqref = 'E2:E10000'

    # Instructions sheet
    wi = wb.create_sheet('Instructions')
    wi.column_dimensions['A'].width = 90
    lines = [
        ('Balance Desk — Import / Export Template', True),
        ('', False),
        ('Sheet tabs', True),
        ('  Each tab represents one month. Name tabs "Mon YY" (e.g. Jun 26, Dec 25).', False),
        ('  You can have multiple month tabs in one file — all will be imported.', False),
        ('', False),
        ('Column guide', True),
        ('  Sr. No.      — Row number (auto-filled on export; ignored on import)', False),
        ('  Date         — YYYY-MM-DD  (e.g. 2026-06-01)', False),
        ('  Title        — Expense name  (required)', False),
        ('  Description  — Optional detail', False),
        ('  Category     — Any text; column E has a preset dropdown', False),
        ('  Mode         — Payment method or payer:', False),
        ('                 Use your card/account name if YOU paid (e.g. AMEX, Chase Visa)', False),
        ('                 Use FRIEND / "Name Paid" if someone ELSE paid (e.g. FRIEND, Aayush Paid)', False),
        ('  Amount       — Full transaction amount  (number only, no currency symbol)', False),
        ('  Split        — Your share of the expense:', False),
        ('                 • Same as Amount → you paid it all for yourself', False),
        ('                 • Less than Amount and YOU paid → friend owes you the difference', False),
        ('                 • Your share and FRIEND paid → what you owe the friend', False),
        ('', False),
        ('Tips', True),
        ('  • Delete the grey example rows before importing.', False),
        ('  • Do not rename or reorder the column headers in row 1.', False),
        ('  • Formulas in Amount/Split cells are supported — Excel evaluates them on save.', False),
    ]
    for r, (text, bold) in enumerate(lines, 1):
        c = wi.cell(r, 1, text)
        c.font = Font(bold=bold, size=11 if bold else 10)
        c.alignment = Alignment(wrap_text=True)

    return wb


# ------------------------------------------------------------------ #
# Routes                                                              #
# ------------------------------------------------------------------ #

@import_export_bp.route('/export-data')
@login_required
def export_data():
    now = datetime.now()
    return render_template('export_data.html',
                           year=now.year,
                           month_names=MONTH_NAMES,
                           current_month=now.month)


@import_export_bp.route('/import-data')
@login_required
def import_data():
    user_id = session['user_id']
    show_normalize = bool(request.args.get('normalize'))
    normalize_modes = normalize_cats = None

    if show_normalize:
        normalize_modes = [r[0] for r in
            db.session.query(Expense.mode)
            .filter_by(user_id=user_id)
            .filter(Expense.mode != None, Expense.mode != '')
            .distinct().order_by(Expense.mode).all()]
        normalize_cats = [r[0] for r in
            db.session.query(Expense.category)
            .filter_by(user_id=user_id)
            .filter(Expense.category != None, Expense.category != '')
            .distinct().order_by(Expense.category).all()]

    return render_template('import_data.html',
                           show_normalize=show_normalize,
                           normalize_modes=normalize_modes,
                           normalize_cats=normalize_cats,
                           standard_categories=DEFAULT_CATEGORIES)


@import_export_bp.route('/download-template')
@login_required
def download_template():
    buf = io.BytesIO()
    _build_template().save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='expenses_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@import_export_bp.route('/export-expenses', methods=['POST'])
@login_required
def export_expenses():
    user_id = session['user_id']
    year    = request.form.get('year', type=int) or datetime.now().year
    raw     = request.form.getlist('months')

    selected = list(range(1, 13)) if (not raw or 'all' in raw) \
               else [int(m) for m in raw if m.isdigit()]
    if not selected:
        flash('Select at least one month.', 'warning')
        return redirect(url_for('import_export.import_export'))

    all_year = (Expense.query
                .filter_by(user_id=user_id)
                .filter(db.extract('year', Expense.date) == year)
                .order_by(Expense.date.asc())
                .all())

    by_month = {}
    for e in all_year:
        if e.date.month in selected:
            key = (e.date.year, e.date.month)
            by_month.setdefault(key, []).append(e)

    if not by_month:
        flash('No expenses found for the selected period.', 'warning')
        return redirect(url_for('import_export.import_export'))

    buf = io.BytesIO()
    _build_export_wb(by_month).save(buf)
    buf.seek(0)

    if len(selected) == 12:
        fname = f'expenses-{year}.xlsx'
    elif len(selected) == 1:
        fname = f'expenses-{_tab_name(year, selected[0]).replace(" ", "-")}.xlsx'
    else:
        fname = f'expenses-{year}-selected.xlsx'

    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@import_export_bp.route('/normalize-data')
@login_required
def normalize_data():
    user_id = session['user_id']
    modes = (db.session.query(Expense.mode)
             .filter_by(user_id=user_id)
             .filter(Expense.mode != None, Expense.mode != '')
             .distinct()
             .order_by(Expense.mode)
             .all())
    categories = (db.session.query(Expense.category)
                  .filter_by(user_id=user_id)
                  .filter(Expense.category != None, Expense.category != '')
                  .distinct()
                  .order_by(Expense.category)
                  .all())
    return {
        'modes': [r[0] for r in modes],
        'categories': [r[0] for r in categories],
        'standard_categories': DEFAULT_CATEGORIES,
    }


@import_export_bp.route('/apply-normalize', methods=['POST'])
@login_required
def apply_normalize():
    user_id = session['user_id']
    changed = 0

    # Mode remaps: form field names are "mode_<old_value>"
    for key, new_val in request.form.items():
        if not key.startswith('mode_'):
            continue
        old_mode = key[5:]          # strip 'mode_' prefix
        new_val  = new_val.strip()
        if not new_val or new_val == old_mode:
            continue
        n = (Expense.query
             .filter_by(user_id=user_id, mode=old_mode)
             .update({'mode': new_val}, synchronize_session=False))
        changed += n

    # Category remaps: form field names are "cat_<old_value>"
    for key, new_val in request.form.items():
        if not key.startswith('cat_'):
            continue
        old_cat = key[4:]
        new_val = new_val.strip()
        if not new_val or new_val == old_cat:
            continue
        n = (Expense.query
             .filter_by(user_id=user_id, category=old_cat)
             .update({'category': new_val}, synchronize_session=False))
        changed += n

    db.session.commit()
    if changed:
        flash(f'Updated {changed} expense{"s" if changed != 1 else ""}. Your data is now consistent.', 'success')
    else:
        flash('No changes were applied.', 'info')
    return redirect(url_for('import_export.import_data'))


# ------------------------------------------------------------------ #
# Import helpers                                                      #
# ------------------------------------------------------------------ #

def _parse_wb(wb):
    """Parse a workbook. Returns (rows_data, nonstandard_cats, skipped, errors).
    rows_data is a list of dicts ready to become Expense objects.
    nonstandard_cats is {original_string: count} for categories not in the standard list.
    """
    rows_data   = []
    nonstandard = {}   # {raw_cat_str: count}
    skipped     = 0
    errors      = []

    for sheet_name in wb.sheetnames:
        parsed = _parse_tab_name(sheet_name)
        if not parsed:
            errors.append(f'Skipped sheet "{sheet_name}" — name not in "Mon YY" format.')
            continue
        sheet_year, sheet_month = parsed
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        layout = _detect_layout(rows[0])
        if layout['date'] < 0 or layout['title'] < 0 or layout['amount'] < 0:
            errors.append(f'Skipped sheet "{sheet_name}" — could not find required columns.')
            continue

        for row in rows[1:]:
            sr_no = _cell(row, 0)
            if not isinstance(sr_no, (int, float)):
                continue
            title  = _cell(row, layout['title'])
            amount = _to_float(_cell(row, layout['amount']))
            if not title or amount is None or amount <= 0:
                skipped += 1
                continue

            raw_date = _cell(row, layout['date'])
            try:
                if isinstance(raw_date, datetime):
                    exp_date = raw_date.date()
                elif isinstance(raw_date, date):
                    exp_date = raw_date
                elif isinstance(raw_date, str):
                    exp_date = date.fromisoformat(raw_date[:10])
                else:
                    raise ValueError
            except (ValueError, TypeError):
                skipped += 1
                continue

            raw_cat  = str(_cell(row, layout['cat'])  or '').strip()
            mode     = str(_cell(row, layout['mode']) or '').strip()
            who_paid = _cell(row, layout['who_paid']) if layout['who_paid'] >= 0 else None
            split_raw = _to_float(_cell(row, layout['split']))
            paid_by_user = _infer_paid_by_user(mode, who_paid)

            if split_raw is None:
                e_split = None
            elif paid_by_user and abs(split_raw - amount) < 0.005:
                e_split = None
            else:
                e_split = round(split_raw, 2)

            # Track non-standard categories (keep raw for user review)
            if raw_cat and raw_cat.lower() not in _STANDARD_CATS_LOWER:
                nonstandard[raw_cat] = nonstandard.get(raw_cat, 0) + 1

            rows_data.append({
                'date':         exp_date,
                'title':        str(title).strip(),
                'description':  str(_cell(row, layout['desc']) or '').strip(),
                'raw_category': raw_cat,
                'mode':         mode,
                'amount':       round(amount, 2),
                'split':        e_split,
                'paid_by_user': paid_by_user,
                'month_key':    (sheet_year, sheet_month),
            })

    return rows_data, nonstandard, skipped, errors


def _commit_import(user_id, rows_data, import_mode, cat_map):
    """Apply cat_map and save rows_data to the DB. Returns (imported, deleted)."""
    seen_months = {r['month_key'] for r in rows_data}
    deleted = 0
    if import_mode == 'replace':
        for yr, mo in seen_months:
            deleted += (Expense.query
                        .filter_by(user_id=user_id)
                        .filter(db.extract('year',  Expense.date) == yr)
                        .filter(db.extract('month', Expense.date) == mo)
                        .delete(synchronize_session=False))

    for r in rows_data:
        raw_cat  = r['raw_category']
        category = cat_map.get(raw_cat) or _normalize_category(raw_cat)
        db.session.add(Expense(
            user_id=user_id, date=r['date'],
            title=r['title'], description=r['description'],
            category=category, mode=r['mode'],
            amount=r['amount'], split=r['split'],
            paid_by_user=r['paid_by_user'],
        ))
    db.session.commit()
    return len(rows_data), deleted


# ------------------------------------------------------------------ #
# Import routes                                                       #
# ------------------------------------------------------------------ #

@import_export_bp.route('/import-expenses', methods=['POST'])
@login_required
def import_expenses():
    import os, uuid, pickle, tempfile

    user_id     = session['user_id']
    import_mode = request.form.get('import_mode', 'add')
    f = request.files.get('file')

    if not f or not f.filename.endswith('.xlsx'):
        flash('Please upload a valid .xlsx file.', 'danger')
        return redirect(url_for('import_export.import_data'))

    # Save uploaded file to a temp path so we can re-use it in step 2
    tmp_dir  = tempfile.gettempdir()
    tmp_name = f'bdimport_{user_id}_{uuid.uuid4().hex}.xlsx'
    tmp_path = os.path.join(tmp_dir, tmp_name)
    f.save(tmp_path)

    try:
        wb = load_workbook(tmp_path, data_only=True)
    except Exception:
        os.unlink(tmp_path)
        flash('Could not read the file. Make sure it is a valid .xlsx.', 'danger')
        return redirect(url_for('import_export.import_data'))

    rows_data, nonstandard, skipped, errors = _parse_wb(wb)

    if not rows_data:
        os.unlink(tmp_path)
        msg = 'No valid expenses found in the file.'
        if skipped:
            msg += f' {skipped} row{"s" if skipped != 1 else ""} were skipped.'
        flash(msg, 'warning')
        for e in errors[:3]:
            flash(e, 'warning')
        return redirect(url_for('import_export.import_data'))

    if nonstandard:
        # Step 2: show category review page
        # Persist parsed data so we don't re-parse on confirm
        data_path = tmp_path.replace('.xlsx', '.pkl')
        with open(data_path, 'wb') as fh:
            pickle.dump({'rows': rows_data, 'skipped': skipped, 'errors': errors}, fh)
        session['import_tmp_xlsx'] = tmp_path
        session['import_tmp_data'] = data_path
        session['import_mode']     = import_mode
        return render_template('import_review.html',
                               nonstandard=sorted(nonstandard.items(), key=lambda x: -x[1]),
                               total_rows=len(rows_data),
                               skipped=skipped,
                               standard_categories=DEFAULT_CATEGORIES)

    # No non-standard categories — import directly
    os.unlink(tmp_path)
    imported, deleted = _commit_import(user_id, rows_data, import_mode, {})
    _flash_import_success(imported, deleted, len({r['month_key'] for r in rows_data}),
                          import_mode, skipped, errors)
    return redirect(url_for('import_export.import_data', normalize=1))


@import_export_bp.route('/import-apply', methods=['POST'])
@login_required
def import_apply():
    import os, pickle

    user_id = session['user_id']
    tmp_xlsx = session.pop('import_tmp_xlsx', None)
    data_path = session.pop('import_tmp_data', None)
    import_mode = session.pop('import_mode', 'add')

    if not data_path or not os.path.exists(data_path):
        flash('Session expired. Please upload the file again.', 'warning')
        return redirect(url_for('import_export.import_data'))

    with open(data_path, 'rb') as fh:
        saved = pickle.load(fh)
    rows_data = saved['rows']
    skipped   = saved['skipped']
    errors    = saved['errors']

    # Build category map from the review form
    cat_map = {}
    for key, val in request.form.items():
        if key.startswith('cat_'):
            raw = key[4:]          # strip 'cat_' prefix
            cat_map[raw] = val.strip()

    imported, deleted = _commit_import(user_id, rows_data, import_mode, cat_map)

    # Clean up temp files
    for path in [tmp_xlsx, data_path]:
        if path and os.path.exists(path):
            try:
                os.unlink(path)
            except OSError:
                pass

    _flash_import_success(imported, deleted, len({r['month_key'] for r in rows_data}),
                          import_mode, skipped, errors)
    return redirect(url_for('import_export.import_data', normalize=1))


def _flash_import_success(imported, deleted, n_months, import_mode, skipped, errors):
    msg = f'Imported {imported} expense{"s" if imported != 1 else ""}.'
    if import_mode == 'replace' and deleted:
        msg += (f' Replaced {deleted} existing expense{"s" if deleted != 1 else ""}'
                f' across {n_months} month{"s" if n_months != 1 else ""}.')
    if skipped:
        msg += f' {skipped} row{"s" if skipped != 1 else ""} skipped.'
    flash(msg, 'success')
    for e in errors[:3]:
        flash(e, 'warning')
